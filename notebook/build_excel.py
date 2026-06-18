"""
Builds the final Excel deliverable with three sheets:

    1. Reconciliation Detail   -- every transaction with its Status/Note
    2. Summary                 -- count of transactions per Status
    3. Reconciliation Statement -- standard Book-to-Bank reconciliation
                                    format, built with live Excel formulas
                                    (not hardcoded Python-calculated values)

Per the project's xlsx conventions: hardcoded inputs are blue, formulas
are black, and the two adjusted balances are highlighted so the reviewer
can see at a glance whether they tie out.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reconcile import load_data, classify_transactions, build_summary

BLUE = Font(color="0000FF")          # hardcoded inputs
BLACK = Font(color="000000")          # formulas
BOLD = Font(bold=True)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HIGHLIGHT_FILL = PatternFill("solid", start_color="FFFF00")
THIN_BORDER = Border(*([Side(style="thin")] * 4))

STATUS_COLORS = {
    "Matched": "C6EFCE",
    "Outstanding Item": "FFEB9C",
    "Amount Mismatch": "FFC7CE",
    "Possible Duplicate": "FFC7CE",
    "Bank Only": "FFD9B3",
    "Book Only": "FFD9B3",
}


def style_header_row(ws, row_idx, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def autosize_columns(ws, df, start_col=1):
    for i, col in enumerate(df.columns):
        if len(df):
            max_len = df[col].apply(lambda x: len(str(x)) if pd.notna(x) else 0).max()
        else:
            max_len = 0
        max_len = max(max_len, len(str(col)))
        ws.column_dimensions[get_column_letter(start_col + i)].width = max_len + 4


def write_detail_sheet(wb, result_df):
    ws = wb.create_sheet("Reconciliation Detail")
    ws.append(list(result_df.columns))
    style_header_row(ws, 1, len(result_df.columns))

    status_col_idx = list(result_df.columns).index("Status") + 1
    amount_bank_idx = list(result_df.columns).index("Amount (Bank)") + 1
    amount_book_idx = list(result_df.columns).index("Amount (Book)") + 1
    NUM_FMT = '#,##0.00;(#,##0.00)'

    for row in result_df.itertuples(index=False):
        ws.append(list(row))
        r = ws.max_row
        status_val = row[status_col_idx - 1]
        fill_color = STATUS_COLORS.get(status_val)
        if fill_color:
            ws.cell(row=r, column=status_col_idx).fill = PatternFill("solid", start_color=fill_color)
        ws.cell(row=r, column=amount_bank_idx).number_format = NUM_FMT
        ws.cell(row=r, column=amount_book_idx).number_format = NUM_FMT
        for c in range(1, len(result_df.columns) + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

    autosize_columns(ws, result_df)
    ws.freeze_panes = "A2"
    return ws


def write_summary_sheet(wb, summary_df, total_transactions):
    ws = wb.create_sheet("Summary")
    ws.append(["Status", "Count"])
    style_header_row(ws, 1, 2)

    start_row = 2
    for row in summary_df.itertuples(index=False):
        ws.append(list(row))
        r = ws.max_row
        fill_color = STATUS_COLORS.get(row.Status)
        if fill_color:
            ws.cell(row=r, column=1).fill = PatternFill("solid", start_color=fill_color)
        ws.cell(row=r, column=2).font = BLACK
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2).border = THIN_BORDER

    end_row = ws.max_row
    total_row = end_row + 2
    ws.cell(row=total_row, column=1, value="Total Transactions").font = BOLD
    # Formula, not a hardcoded Python value -- recalculates if the detail changes
    ws.cell(row=total_row, column=2, value=f"=SUM(B{start_row}:B{end_row})").font = BLACK
    ws.cell(row=total_row, column=2).font = BOLD

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    return ws


def write_reconciliation_statement_sheet(wb):
    """
    Builds the classic two-column Bank Reconciliation Statement.

    The standard logic (Gleim/AICPA-style "two-balance" format):

        Balance per Bank Statement
          + Deposits in Transit       (recorded in books, not yet by bank)
          - Outstanding Checks        (recorded in books, not yet by bank)
        = Adjusted Bank Balance

        Balance per Books
          + Bank credits not yet recorded in books (e.g. interest earned)
          - Bank charges not yet recorded in books (e.g. service fees)
        = Adjusted Book Balance

    Adjusted Bank Balance must equal Adjusted Book Balance.

    Mapping to this project's Status categories:
        - "Book Only" rows are transactions the company has recorded but
          the bank hasn't cleared yet -> these adjust the BANK balance
          (Deposits in Transit if positive, Outstanding Checks if negative).
        - "Bank Only" rows are transactions the bank has recorded but the
          company hasn't booked yet (e.g. bank fees) -> these adjust the
          BOOK balance.

    Every figure is an Excel formula referencing the Reconciliation Detail
    sheet, so changing the underlying data automatically updates this
    statement -- nothing here is a hardcoded Python-calculated number.
    """
    ws = wb.create_sheet("Reconciliation Statement", 0)  # put it first
    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 16
    NUM_FMT = '#,##0.00;(#,##0.00)'

    title = ws.cell(row=1, column=1, value="Bank Reconciliation Statement")
    title.font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="As of period end (see Reconciliation Detail tab for source data)").font = Font(italic=True, size=9)

    detail_ref = "'Reconciliation Detail'"

    # ---- Bank side (left column) ----
    r = 4
    ws.cell(row=r, column=1, value="BALANCE PER BANK STATEMENT").font = BOLD
    r += 1
    ws.cell(row=r, column=1, value="Balance per bank statement, ending")
    c = ws.cell(row=r, column=2, value=f"=SUM({detail_ref}!F:F)")
    c.font = BLACK
    c.number_format = NUM_FMT
    bank_balance_row = r
    r += 1
    ws.cell(row=r, column=1, value="(+) Deposits in Transit (recorded in books, not yet by bank)").alignment = Alignment(wrap_text=True)
    c = ws.cell(row=r, column=2,
            value=f"=SUMIFS({detail_ref}!G:G,{detail_ref}!H:H,\"Book Only\",{detail_ref}!G:G,\">0\")")
    c.font = BLACK
    c.number_format = NUM_FMT
    dit_row = r
    r += 1
    ws.cell(row=r, column=1, value="(+) Outstanding Checks (recorded in books, not yet by bank)").alignment = Alignment(wrap_text=True)
    c = ws.cell(row=r, column=2,
            value=f"=SUMIFS({detail_ref}!G:G,{detail_ref}!H:H,\"Book Only\",{detail_ref}!G:G,\"<0\")")
    c.font = BLACK
    c.number_format = NUM_FMT
    oc_row = r
    r += 2
    ws.cell(row=r, column=1, value="Adjusted Bank Balance").font = BOLD
    # Outstanding checks are already negative amounts, so adding them
    # back reduces the balance correctly -- no separate subtraction needed.
    c = ws.cell(row=r, column=2,
            value=f"=B{bank_balance_row}+B{dit_row}+B{oc_row}")
    c.font = BOLD
    c.number_format = NUM_FMT
    adj_bank_row = r
    for cc in range(1, 3):
        ws.cell(row=adj_bank_row, column=cc).fill = HIGHLIGHT_FILL
        ws.cell(row=adj_bank_row, column=cc).border = THIN_BORDER

    # ---- Book side (right column, starting at column D) ----
    r = 4
    ws.cell(row=r, column=4, value="BALANCE PER BOOKS").font = BOLD
    r += 1
    ws.cell(row=r, column=4, value="Balance per books, ending")
    c = ws.cell(row=r, column=5, value=f"=SUM({detail_ref}!G:G)")
    c.font = BLACK
    c.number_format = NUM_FMT
    book_balance_row = r
    r += 1
    ws.cell(row=r, column=4, value="(+/-) Bank-side items not yet recorded in books (e.g. fees)").alignment = Alignment(wrap_text=True)
    c = ws.cell(row=r, column=5,
            value=f"=SUMIFS({detail_ref}!F:F,{detail_ref}!H:H,\"Bank Only\")")
    c.font = BLACK
    c.number_format = NUM_FMT
    bank_only_row = r
    r += 2
    ws.cell(row=r, column=4, value="Adjusted Book Balance").font = BOLD
    c = ws.cell(row=r, column=5, value=f"=E{book_balance_row}+E{bank_only_row}")
    c.font = BOLD
    c.number_format = NUM_FMT
    adj_book_row = r
    for cc in range(4, 6):
        ws.cell(row=adj_book_row, column=cc).fill = HIGHLIGHT_FILL
        ws.cell(row=adj_book_row, column=cc).border = THIN_BORDER

    # ---- Tie-out check ----
    tie_row = max(adj_book_row, adj_bank_row) + 3
    ws.cell(row=tie_row, column=1, value="Difference (Adjusted Bank - Adjusted Book)").font = BOLD
    c = ws.cell(row=tie_row, column=2, value=f"=B{adj_bank_row}-E{adj_book_row}")
    c.font = BLACK
    c.number_format = NUM_FMT
    ws.cell(row=tie_row + 1, column=1, value="Reconciled?").font = BOLD
    ws.cell(
        row=tie_row + 1, column=2,
        value=f'=IF(ROUND(B{tie_row},2)=0,"YES - Tied Out","NO - Review Required")'
    ).font = BOLD

    # ---- Explain the difference: unresolved items not included in either
    #      adjustment above (Amount Mismatch and Possible Duplicate items
    #      are deliberately NOT auto-included, since a human needs to
    #      decide how to treat them -- see note below) ----
    explain_row = tie_row + 3
    ws.cell(row=explain_row, column=1, value="Reconciling difference is explained by unresolved items:").font = Font(italic=True, bold=True, size=9)
    r2 = explain_row + 1
    ws.cell(row=r2, column=1, value="Amount Mismatch -- bank amount not yet corrected in books").alignment = Alignment(wrap_text=True)
    c = ws.cell(row=r2, column=2,
            value=f"=SUMIFS({detail_ref}!F:F,{detail_ref}!H:H,\"Amount Mismatch\")-SUMIFS({detail_ref}!G:G,{detail_ref}!H:H,\"Amount Mismatch\")")
    c.font = BLACK
    c.number_format = NUM_FMT
    r2 += 1
    ws.cell(row=r2, column=1, value="Possible Duplicate -- bank amount with no offsetting book entry").alignment = Alignment(wrap_text=True)
    c = ws.cell(row=r2, column=2,
            value=(f"=SUMIFS({detail_ref}!F:F,{detail_ref}!H:H,\"Possible Duplicate\")"
                   f"-SUMIFS({detail_ref}!G:G,{detail_ref}!H:H,\"Possible Duplicate\")"))
    c.font = BLACK
    c.number_format = NUM_FMT
    r2 += 1
    sum_explain_end = r2 - 1
    ws.cell(row=r2, column=1, value="Sum of unresolved items (should equal Difference above)").font = BOLD
    c = ws.cell(row=r2, column=2, value=f"=SUM(B{explain_row+1}:B{sum_explain_end})")
    c.font = BOLD
    c.number_format = NUM_FMT

    note_row = r2 + 2
    ws.cell(row=note_row, column=1,
            value=("Note: This is a simplified statement built from the sample dataset. Adjusted Bank "
                   "and Adjusted Book balances will only tie out once Amount Mismatch and Possible "
                   "Duplicate items are resolved by a human reviewer (corrected, reversed, or confirmed "
                   "as legitimate) -- they are intentionally excluded from the automatic Deposits in "
                   "Transit / Outstanding Checks / Bank-only adjustments above so they are never silently "
                   "netted out without review. Outstanding Item (pure date timing differences with no "
                   "amount issue) does not affect either ending balance and is excluded for the same reason.")
            ).font = Font(italic=True, size=8, color="808080")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    ws.row_dimensions[note_row].height = 70
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    # Wrapped-text rows need explicit height or LibreOffice/Excel will clip them
    for wrapped_row in (dit_row, oc_row, bank_only_row, explain_row + 1, explain_row + 2):
        ws.row_dimensions[wrapped_row].height = 30

    return ws


def main():
    bank, book = load_data(
        "/home/claude/AI-Bank-Reconciliation/data/bank_transactions.csv",
        "/home/claude/AI-Bank-Reconciliation/data/book_transactions.csv",
    )
    result_df = classify_transactions(bank, book)
    summary_df = build_summary(result_df)

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    write_detail_sheet(wb, result_df)
    write_summary_sheet(wb, summary_df, len(result_df))
    write_reconciliation_statement_sheet(wb)  # inserted at index 0 (first tab)

    wb.active = 0
    output_path = "/home/claude/AI-Bank-Reconciliation/output/reconciliation_output.xlsx"
    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
